# Server_Core/app/api/routes.py
import os
import httpx
import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import or_, cast, String, text, func, case
from fastapi import APIRouter, Depends, HTTPException, Request, WebSocket, WebSocketDisconnect
import base64
import cv2
from fastapi import WebSocket, WebSocketDisconnect
import json
import asyncio
import numpy as np
from app.ai.engine import face_engine 
from pydantic import BaseModel
from sqlalchemy.dialects.postgresql import insert
from app.db.session import get_db
from app.db.models import *
import io
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from fastapi.responses import StreamingResponse, HTMLResponse


router = APIRouter()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Định nghĩa Schema cho Request gửi lên
class FaceEnrollRequest(BaseModel):
    sv_id: int
    gv_id: int
    images: list[str]

# Định nghĩa viền bảng chung
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def model_to_dict(model_obj):
    return {c.name: getattr(model_obj, c.name) for c in model_obj.__table__.columns}

# ---------------------------------------------------------
# API ĐĂNG NHẬP (Proxy chuyển tiếp lên Supabase)
# ---------------------------------------------------------
@router.post("/auth/v1/token")
async def login_proxy(request: Request):
    """Giả lập endpoint đăng nhập của Supabase để App không bị lỗi"""
    body = await request.json()
    grant_type = request.query_params.get("grant_type", "password")
    
    async with httpx.AsyncClient() as client:
        res = await client.post(
            f"{SUPABASE_URL}/auth/v1/token?grant_type={grant_type}",
            json=body,
            headers={"apikey": SUPABASE_KEY, "Content-Type": "application/json"}
        )
        # Trả về y nguyên kết quả (kể cả lỗi sai mật khẩu hay thành công)
        if res.status_code != 200:
            raise HTTPException(status_code=res.status_code, detail=res.json())
        return res.json()

# ---------------------------------------------------------
# API QUẢN LÝ GIẢNG VIÊN (Lồng tên Khoa)
# ---------------------------------------------------------
@router.get("/giangvien")
async def get_giangvien(id: str = None, auth_id: str = None, db: AsyncSession = Depends(get_db)):
    stmt = select(GiangVien, Khoa).outerjoin(Khoa, GiangVien.khoa_id == Khoa.id)
    
    # Client tìm theo ID (Ví dụ: id=eq.3)
    if id and id.startswith("eq."):
        gv_id = int(id.replace("eq.", ""))
        stmt = stmt.where(GiangVien.id == gv_id)
        
    # Client tìm theo Auth UUID sau khi đăng nhập
    if auth_id and auth_id.startswith("eq."):
        a_id = auth_id.replace("eq.", "")
        stmt = stmt.where(GiangVien.auth_id == a_id)
        
    result = await db.execute(stmt)
    res = []
    for gv, khoa in result:
        d = model_to_dict(gv)
        d["created_at"] = str(d["created_at"]) if d.get("created_at") else None
        d["khoa"] = {"tenkhoa": khoa.tenkhoa} if khoa else None
        res.append(d)
    return res

# ---------------------------------------------------------
# API TUẦN HỌC
# ---------------------------------------------------------
@router.get("/tuan_hoc")
async def get_tuan_hoc(db: AsyncSession = Depends(get_db)):
    stmt = select(TuanHoc).order_by(TuanHoc.id.asc())
    result = await db.execute(stmt)
    res = []
    for t in result.scalars().all():
        d = model_to_dict(t)
        d["ngay_bat_dau"] = str(d["ngay_bat_dau"]) if d.get("ngay_bat_dau") else None
        d["ngay_ket_thuc"] = str(d["ngay_ket_thuc"]) if d.get("ngay_ket_thuc") else None
        res.append(d)
    return res

# ---------------------------------------------------------
# API SINH VIÊN
# ---------------------------------------------------------
@router.get("/sinhvien")
async def get_danh_sach_sinh_vien(class_id: str = None, db: AsyncSession = Depends(get_db)):
    stmt = select(SinhVien)
    if class_id and class_id.startswith("eq."):
        c_id = class_id.replace("eq.", "")
        stmt = stmt.where(SinhVien.class_id == c_id)
    result = await db.execute(stmt)
    res = []
    for sv in result.scalars().all():
        d = model_to_dict(sv)
        d["ngaysinh"] = str(d["ngaysinh"]) if d.get("ngaysinh") else None
        d["created_at"] = str(d["created_at"]) if d.get("created_at") else None
        res.append(d)
    return res

# ---------------------------------------------------------
# CÁC API CŨ LẤY THÔNG BÁO
# ---------------------------------------------------------
@router.get("/thongbao")
async def get_thongbao(limit: int = 3, db: AsyncSession = Depends(get_db)):
    stmt = select(ThongBao).order_by(ThongBao.created_at.desc()).limit(limit)
    result = await db.execute(stmt)
    res = []
    for t in result.scalars().all():
        d = model_to_dict(t)
        d["created_at"] = str(t.created_at) if t.created_at else None
        res.append(d)
    return res

# ---------------------------------------------------------
# API THỜI KHÓA BIỂU
# ---------------------------------------------------------
@router.get("/thoikhoabieu")
async def get_thoi_khoa_bieu(giangvien_id: str = None, db: AsyncSession = Depends(get_db)):
    stmt = select(ThoiKhoaBieu, Lop, HocPhan, HocKy)\
        .outerjoin(Lop, ThoiKhoaBieu.lop_id == Lop.id)\
        .outerjoin(HocPhan, ThoiKhoaBieu.hocphan_id == HocPhan.id)\
        .outerjoin(HocKy, ThoiKhoaBieu.hocky_id == HocKy.id)
        
    if giangvien_id and giangvien_id.startswith("eq."):
        gv_id = int(giangvien_id.replace("eq.", ""))
        stmt = stmt.where(ThoiKhoaBieu.giangvien_id == gv_id)
        
    result = await db.execute(stmt)
    res = []
    for tkb, lop, hocphan, hocky in result:
        # Lấy TOÀN BỘ các trường gốc của bảng thoikhoabieu (kể cả hocphan_id, lop_id...)
        d = model_to_dict(tkb)
        
        # Nhồi thêm các Object lồng nhau vào
        d["lop"] = {"tenlop": lop.tenlop} if lop else None
        d["hocphan"] = {"tenhocphan": hocphan.tenhocphan, "sobuoi": hocphan.sobuoi} if hocphan else None
        d["hocky"] = {"namhoc": hocky.namhoc, "tenhocky": hocky.tenhocky} if hocky else None
        
        res.append(d)
    return res

# ---------------------------------------------------------
# API TIẾT HỌC
# ---------------------------------------------------------
@router.get("/tkb_tiet")
async def get_tkb_tiet(tkb_id: str = None, thu: str = None, db: AsyncSession = Depends(get_db)):
    stmt = select(TKBTiet, Tiet).outerjoin(Tiet, TKBTiet.tiet_id == Tiet.id)
    
    if tkb_id:
        if tkb_id.startswith("in."):
            ids_str = tkb_id.replace("in.(", "").replace(")", "")
            ids = [int(i) for i in ids_str.split(",") if i.strip()]
            stmt = stmt.where(TKBTiet.tkb_id.in_(ids))
        # --- BỔ SUNG NHÁNH NÀY ---
        elif tkb_id.startswith("eq."):
            id_val = int(tkb_id.replace("eq.", ""))
            stmt = stmt.where(TKBTiet.tkb_id == id_val)
            
    if thu and thu.startswith("eq."):
        thu_val = int(thu.replace("eq.", ""))
        stmt = stmt.where(TKBTiet.thu == thu_val)
        
    stmt = stmt.order_by(Tiet.thoigianbd.asc())
        
    result = await db.execute(stmt)
    res = []
    for tkbtiet, tiet in result:
        d = model_to_dict(tkbtiet)
        d["created_at"] = str(d["created_at"]) if d.get("created_at") else None
        d["tiet"] = {
            "thoigianbd": str(tiet.thoigianbd) if tiet.thoigianbd else None, 
            "thoigiankt": str(tiet.thoigiankt) if tiet.thoigiankt else None
        } if tiet else None
        res.append(d)
    return res

# ---------------------------------------------------------
# API ĐIỂM DANH
# ---------------------------------------------------------
@router.get("/diemdanh")
async def get_diemdanh(tkb_tiet_id: str = None, ngay_diem_danh: str = None, db: AsyncSession = Depends(get_db)):
    stmt = select(DiemDanh)
    
    if tkb_tiet_id:
        if tkb_tiet_id.startswith("in."):
            ids_str = tkb_tiet_id.replace("in.(", "").replace(")", "")
            ids = [int(i) for i in ids_str.split(",") if i.strip()]
            stmt = stmt.where(DiemDanh.tkb_tiet_id.in_(ids))
        # --- BỔ SUNG NHÁNH NÀY ---
        elif tkb_tiet_id.startswith("eq."):
            id_val = int(tkb_tiet_id.replace("eq.", ""))
            stmt = stmt.where(DiemDanh.tkb_tiet_id == id_val)
            
    if ngay_diem_danh and ngay_diem_danh.startswith("eq."):
        date_str = ngay_diem_danh.replace("eq.", "")
        target_date = datetime.date.fromisoformat(date_str)
        stmt = stmt.where(DiemDanh.ngay_diem_danh == target_date)
        
    result = await db.execute(stmt)
    res = []
    for d in result.scalars().all():
        d_dict = model_to_dict(d)
        d_dict["ngay_diem_danh"] = str(d.ngay_diem_danh) if d.ngay_diem_danh else None
        d_dict["created_at"] = str(d.created_at) if d.created_at else None
        res.append(d_dict)
    return res

#===========================================================
# API NHẬN DẠNG ĐIỂM DANH
#==========================================================
@router.websocket("/ws/attendance/{tkb_tiet_id}")
async def attendance_websocket(websocket: WebSocket, tkb_tiet_id: int, db: AsyncSession = Depends(get_db)):
    await websocket.accept()
    print(f"[Server] WebSocket kết nối cho Tiết: {tkb_tiet_id}")
    
    session_scanned_ids = set()
    
    try:
        while True:
            data = await websocket.receive_text()
            payload = json.loads(data)
            
            image_b64 = payload.get("image")
            mode = payload.get("mode", "1")
            
            # BẮT LẤY NGÀY MÀ CLIENT GỬI LÊN ĐỂ LƯU CHO CHUẨN
            target_date_str = payload.get("date")
            if target_date_str:
                target_date = datetime.date.fromisoformat(target_date_str)
            else:
                target_date = datetime.date.today()
            
            if not image_b64: continue

            embeddings = await asyncio.to_thread(face_engine.process_attendance_frame, image_b64, mode)
            recognized_students = []
            
            for emb in embeddings:
                stmt = select(FaceEmbedding).order_by(FaceEmbedding.embedding.cosine_distance(emb)).limit(1)
                result = await db.execute(stmt)
                match = result.scalar_one_or_none()

                if match:
                    sv_id = match.sv_id
                    
                    if sv_id in session_scanned_ids:
                        continue
                        
                    sv_stmt = select(SinhVien).where(SinhVien.id == sv_id)
                    sv_res = await db.execute(sv_stmt)
                    sv = sv_res.scalar_one_or_none()
                    
                    if sv:
                        # KIỂM TRA TRÙNG LẶP THEO NGÀY ĐƯỢC CHỌN
                        check_stmt = select(DiemDanh).where(
                            DiemDanh.sv_id == sv_id,
                            DiemDanh.tkb_tiet_id == tkb_tiet_id,
                            DiemDanh.ngay_diem_danh == target_date 
                        )
                        check_res = await db.execute(check_stmt)
                        existing_record = check_res.scalar_one_or_none()
                        
                        ho_ten = f"{sv.hodem} {sv.ten}".strip() if hasattr(sv, 'hodem') else sv.hoten
                        
                        if not existing_record:
                            try:
                                # LƯU VÀO DB BẰNG NGÀY ĐƯỢC CHỌN
                                stmt_insert = insert(DiemDanh).values(
                                    sv_id=sv.id,
                                    tkb_tiet_id=tkb_tiet_id,
                                    ngay_diem_danh=target_date, 
                                    trang_thai="Có mặt"
                                )
                                await db.execute(stmt_insert)
                                await db.commit()
                            except Exception as db_err:
                                print(f"[Lỗi DB] {db_err}")
                                await db.rollback()
                                continue 
                                
                        session_scanned_ids.add(sv_id)
                        
                        recognized_students.append({
                            "id": str(sv.id),
                            "name": ho_ten,
                            "time": datetime.datetime.now().strftime("%H:%M:%S")
                        })

            if recognized_students:
                await websocket.send_text(json.dumps({
                    "status": "success",
                    "students": recognized_students
                }))

    except WebSocketDisconnect:
        print(f"[Server] WebSocket ngắt kết nối: {tkb_tiet_id}")
    except Exception as e:
        print(f"[Server WebSocket Error] {e}")

        
# =========================================================
# API DÀNH RIÊNG CHO TRANG ĐÀO TẠO KHUÔN MẶT (FACE TRAINING)
# =========================================================

from sqlalchemy import or_
from app.db.models import FaceEmbedding

# 1. Lấy danh sách các lớp mà Giảng viên phụ trách
@router.get("/training/giangvien/{gv_id}/lophoc")
async def get_lop_giang_day(gv_id: int, db: AsyncSession = Depends(get_db)):
    """
    Truy vấn bảng ThoiKhoaBieu để tìm các lớp giảng viên đang dạy.
    Sử dụng distinct() để không bị trùng lặp lớp nếu dạy nhiều môn.
    """
    stmt = select(Lop).join(
        ThoiKhoaBieu, Lop.id == ThoiKhoaBieu.lop_id
    ).where(
        ThoiKhoaBieu.giangvien_id == gv_id,
        ThoiKhoaBieu.deleted_at.is_(None) # Bỏ qua các TKB đã bị xóa mềm
    ).distinct()
    
    result = await db.execute(stmt)
    res = []
    for lop in result.scalars().all():
        res.append({
            "id": lop.id,
            "name": lop.tenlop,
            "siso": 0 # TODO: Anh có thể count sinh viên sau nếu cần hiển thị sĩ số
        })
    return res

# 2. Lấy danh sách Sinh viên trong 1 lớp (Kèm trạng thái đã có khuôn mặt chưa)
@router.get("/training/lop/{class_id}/sinhvien")
async def get_sinhvien_training(class_id: str, db: AsyncSession = Depends(get_db)):
    """
    Lấy toàn bộ sinh viên của lớp, LEFT JOIN với FaceEmbedding 
    để biết bạn nào đã có dữ liệu nhận diện.
    """
    stmt = select(SinhVien, FaceEmbedding.sv_id).outerjoin(
        FaceEmbedding, SinhVien.id == FaceEmbedding.sv_id
    ).where(
        SinhVien.class_id == class_id,
        SinhVien.deleted_at.is_(None) # Chỉ lấy sinh viên đang học (chưa bị xóa)
    )
    
    result = await db.execute(stmt)
    res = []
    for sv, face_id in result:
        res.append({
            "id": sv.id,
            "name": f"{sv.hodem} {sv.ten}".strip(),
            "has_data": bool(face_id is not None) # True nếu tìm thấy trong bảng FaceEmbedding
        })
    return res

# 3. Tìm kiếm Sinh viên theo Mã SV hoặc Tên (Trong phạm vi các lớp GV quản lý)
@router.get("/training/giangvien/{gv_id}/timkiem")
async def search_sinhvien(gv_id: int, keyword: str, db: AsyncSession = Depends(get_db)):
    """
    Tìm sinh viên theo từ khóa (Mã SV hoặc Tên) nhưng phải nằm trong 
    danh sách các lớp mà Giảng viên này có dạy.
    """
    # Bước 3.1: Lấy danh sách lớp GV dạy trước
    lop_stmt = select(ThoiKhoaBieu.lop_id).where(
        ThoiKhoaBieu.giangvien_id == gv_id,
        ThoiKhoaBieu.deleted_at.is_(None)
    ).distinct()
    lop_result = await db.execute(lop_stmt)
    lop_ids = [row for row in lop_result.scalars().all()]
    
    if not lop_ids:
        return [] # Nếu GV không dạy lớp nào thì trả về rỗng
        
    # Bước 3.2: Tìm kiếm sinh viên nằm trong các lớp đó
    search_keyword = f"%{keyword}%"
    
    # SỬA LỖI 500 Ở ĐÂY: Dùng cast() của ORM thay vì text() để chuyển Integer sang String
    stmt = select(SinhVien, FaceEmbedding.sv_id).outerjoin(
        FaceEmbedding, SinhVien.id == FaceEmbedding.sv_id
    ).where(
        SinhVien.class_id.in_(lop_ids),
        SinhVien.deleted_at.is_(None),
        or_(
            cast(SinhVien.id, String).ilike(search_keyword), # Ép kiểu ID thành chuỗi để tìm kiếm
            SinhVien.ten.ilike(search_keyword),
            SinhVien.hodem.ilike(search_keyword)
        )
    )
    
    result = await db.execute(stmt)
    res = []
    for sv, face_id in result:
        res.append({
            "id": sv.id,
            "class_id": sv.class_id,
            "name": f"{sv.hodem} {sv.ten}".strip(),
            "has_data": bool(face_id is not None)
        })
    return res

#===============================================
# API ĐÂÒ TẠO DỮ LIỆU
#===============================================
@router.post("/training/face/enroll")
async def enroll_face_data(req: FaceEnrollRequest, db: AsyncSession = Depends(get_db)):
    try:
        # 1. Xử lý qua AI Engine lấy Vector trung bình
        fused_embedding = face_engine.extract_fused_embedding(req.images)
        
        # 2. Xây dựng câu lệnh UPSERT (Cập nhật nếu có, Thêm mới nếu chưa) bằng PostgreSQL
        stmt = insert(FaceEmbedding).values(
            sv_id=req.sv_id,
            embedding=fused_embedding,
            trained_by=req.gv_id
        )
        
        # ON CONFLICT DO UPDATE: Khớp khóa chính sv_id thì ghi đè
        stmt = stmt.on_conflict_do_update(
            index_elements=['sv_id'],
            set_={
                'embedding': fused_embedding,
                'trained_by': req.gv_id,
                'updated_at': text('now()')
            }
        )
        
        # Thực thi lưu DB
        await db.execute(stmt)
        await db.commit()
        
        return {"status": "success", "message": "Cập nhật dữ liệu khuôn mặt thành công!"}
        
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        print(f"[API Enroll Lỗi] {e}")
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi xử lý dữ liệu khuôn mặt.")
    
# ----------------------------------------------------------------------
# API: BÁO CÁO CHI TIẾT ĐIỂM DANH THEO LỚP
# ----------------------------------------------------------------------
@router.get("/export/report/detailed/{tkb_id}")
async def export_detailed_attendance(tkb_id: int, db: AsyncSession = Depends(get_db)):
    # Lấy thông tin TKB và Header
    stmt = select(ThoiKhoaBieu, Lop, HocPhan, HocKy, GiangVien)\
        .join(Lop, ThoiKhoaBieu.lop_id == Lop.id)\
        .join(HocPhan, ThoiKhoaBieu.hocphan_id == HocPhan.id)\
        .join(HocKy, ThoiKhoaBieu.hocky_id == HocKy.id)\
        .join(GiangVien, ThoiKhoaBieu.giangvien_id == GiangVien.id)\
        .where(ThoiKhoaBieu.id == tkb_id)
    
    res = (await db.execute(stmt)).first()
    if not res: raise HTTPException(status_code=404, detail="Không tìm thấy lịch học")
    tkb, lop, hp, hk, gv = res

    wb = load_workbook("app/templates/temp_DetailedAttendance.xlsx")
    ws = wb.active
    
    # Điền Header (Theo mẫu của Bé mèo nhỏ: Cột C, dòng 5-9)
    ws['C5'], ws['C6'], ws['C7'], ws['C8'], ws['C9'] = lop.tenlop, hp.tenhocphan, hk.tenhocky, hk.namhoc, f"{gv.hodem} {gv.ten}"

    # Lấy danh sách ngày đã điểm danh của tkb_id này
    date_stmt = select(DiemDanh.ngay_diem_danh).where(DiemDanh.tkb_tiet_id.in_(
        select(TKBTiet.id).where(TKBTiet.tkb_id == tkb_id)
    )).distinct().order_by(DiemDanh.ngay_diem_danh.asc())
    dates = (await db.execute(date_stmt)).scalars().all()

    # Đổ tiêu đề ngày vào dòng 12 (Từ cột E)
    for i, d in enumerate(dates):
        cell = ws.cell(row=12, column=5+i, value=d.strftime("%d/%m"))
        cell.font, cell.border, cell.alignment = Font(bold=True), thin_border, Alignment(horizontal="center")

    # Đổ danh sách sinh viên (Dòng 13 trở đi)
    sv_stmt = select(SinhVien).where(SinhVien.class_id == lop.id).order_by(SinhVien.id.asc())
    students = (await db.execute(sv_stmt)).scalars().all()

    for r_idx, sv in enumerate(students):
        row = 13 + r_idx
        ws.cell(row=row, column=1, value=r_idx + 1).border = thin_border
        ws.cell(row=row, column=2, value=sv.id).border = thin_border
        ws.cell(row=row, column=3, value=sv.hodem).border = thin_border
        ws.cell(row=row, column=4, value=sv.ten).border = thin_border
        
        # Lấy điểm danh của SV này trong tkb_id hiện tại
        dd_res = await db.execute(select(DiemDanh).where(DiemDanh.sv_id == sv.id, DiemDanh.tkb_tiet_id.in_(
            select(TKBTiet.id).where(TKBTiet.tkb_id == tkb_id)
        )))
        sv_dd_map = {d.ngay_diem_danh: d.trang_thai for d in dd_res.scalars().all()}
        
        for c_idx, d in enumerate(dates):
            status = sv_dd_map.get(d, "")
            cell = ws.cell(row=row, column=5+c_idx, value="X" if status=="Có mặt" else "V" if status=="Vắng" else "")
            cell.border, cell.alignment = thin_border, Alignment(horizontal="center")

    return send_excel_response(wb, f"ChiTietDiemDanh_{lop.id}")

# ----------------------------------------------------------------------
# API: BÁO CÁO TỔNG QUAN PHỤ TRÁCH LỚP
# ----------------------------------------------------------------------
@router.get("/export/report/overview/{gv_id}")
async def export_overview_report(gv_id: int, db: AsyncSession = Depends(get_db)):
    gv_res = await db.execute(select(GiangVien).where(GiangVien.id == gv_id))
    gv = gv_res.scalar_one_or_none()

    wb = load_workbook("app/templates/temp_OverviewReport.xlsx")
    ws = wb.active
    ws['C7'] = f"{gv.hodem} {gv.ten}" if gv else "N/A"

    stmt = select(Lop.id, Lop.tenlop, func.count(SinhVien.id).label("siso"))\
        .join(ThoiKhoaBieu, Lop.id == ThoiKhoaBieu.lop_id)\
        .join(SinhVien, Lop.id == SinhVien.class_id)\
        .where(ThoiKhoaBieu.giangvien_id == gv_id).group_by(Lop.id, Lop.tenlop)
    
    data = (await db.execute(stmt)).all()

    for r_idx, item in enumerate(data):
        row = 10 + r_idx
        tiet_ids_subquery = select(TKBTiet.id).join(ThoiKhoaBieu).where(ThoiKhoaBieu.lop_id == item.id)
        
        # Số buổi đã dạy (đếm số ngày khác nhau)
        buoi_day = await db.scalar(select(func.count(func.distinct(DiemDanh.ngay_diem_danh))).where(DiemDanh.tkb_tiet_id.in_(tiet_ids_subquery))) or 0
        
        # Tổng số lượt có mặt của cả lớp
        tong_co_mat = await db.scalar(select(func.count(DiemDanh.id)).where(DiemDanh.tkb_tiet_id.in_(tiet_ids_subquery))) or 0
        
        # Tỷ lệ chuyên cần = (Tổng lượt có mặt / (Sĩ số * Số buổi)) * 100
        tong_luot_toi_da = item.siso * buoi_day
        ty_le = (tong_co_mat / tong_luot_toi_da * 100) if tong_luot_toi_da > 0 else 0

        vals = [r_idx+1, item.id, item.tenlop, item.siso, buoi_day, f"{ty_le:.1f}%"]
        for c_idx, val in enumerate(vals):
            cell = ws.cell(row=row, column=1+c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    return send_excel_response(wb, f"TongQuanLop_GV_{gv_id}")

# ----------------------------------------------------------------------
# API: DANH SÁCH CẢNH BÁO HỌC VỤ
# ----------------------------------------------------------------------
@router.get("/export/report/warning/{tkb_id}")
async def export_warning_report(tkb_id: int, db: AsyncSession = Depends(get_db)):
    # 1. Lấy thông tin TKB kèm Loại học phần
    stmt = select(ThoiKhoaBieu, Lop, HocPhan, LoaiHocPhan, HocKy, GiangVien)\
        .join(Lop, ThoiKhoaBieu.lop_id == Lop.id)\
        .join(HocPhan, ThoiKhoaBieu.hocphan_id == HocPhan.id)\
        .join(LoaiHocPhan, HocPhan.loaihp_id == LoaiHocPhan.id)\
        .join(HocKy, ThoiKhoaBieu.hocky_id == HocKy.id)\
        .join(GiangVien, ThoiKhoaBieu.giangvien_id == GiangVien.id)\
        .where(ThoiKhoaBieu.id == tkb_id)
        
    res = (await db.execute(stmt)).first()
    if not res: raise HTTPException(status_code=404, detail="Không tìm thấy lịch học")
    tkb, lop, hp, lhp, hk, gv = res

    # 2. Xác định định mức cảnh báo
    # Nếu là Lý thuyết (ID=1 chẳng hạn) -> 30%, còn lại (Thực hành) -> 20%
    threshold_pct = 0.3 if "Lý thuyết" in lhp.tenloai else 0.2
    max_allowed_absences = hp.sobuoi * threshold_pct

    # 3. Lấy số buổi ĐÃ DẠY thực tế đến hiện tại
    tiet_ids = select(TKBTiet.id).where(TKBTiet.tkb_id == tkb_id)
    total_conducted = await db.scalar(select(func.count(func.distinct(DiemDanh.ngay_diem_danh))).where(DiemDanh.tkb_tiet_id.in_(tiet_ids))) or 0

    wb = load_workbook("app/templates/temp_WarningList.xlsx")
    ws = wb.active
    ws['C5'], ws['C6'], ws['C7'], ws['C8'], ws['C9'] = lop.tenlop, hp.tenhocphan, hk.tenhocky, hk.namhoc, f"{gv.hodem} {gv.ten}"

    # 4. Lấy danh sách TẤT CẢ sinh viên trong lớp để kiểm tra
    sv_stmt = select(SinhVien).where(SinhVien.class_id == lop.id).order_by(SinhVien.id.asc())
    students = (await db.execute(sv_stmt)).scalars().all()

    red_font = Font(color="FF0000", bold=True)

    for r_idx, sv in enumerate(students):
        row = 12 + r_idx
        # Đếm số buổi có mặt của SV này
        attended = await db.scalar(select(func.count(DiemDanh.id)).where(
            DiemDanh.sv_id == sv.id, 
            DiemDanh.trang_thai == 'Có mặt',
            DiemDanh.tkb_tiet_id.in_(tiet_ids)
        )) or 0
        
        absent = total_conducted - attended
        is_warning = absent > max_allowed_absences
        
        percentage = (absent / hp.sobuoi * 100) if hp.sobuoi > 0 else 0
        status_text = "CẢNH BÁO" if is_warning else "Bình thường"

        vals = [r_idx+1, sv.id, sv.hodem, sv.ten, attended, absent, f"{percentage:.1f}%", status_text]
        for c_idx, val in enumerate(vals):
            cell = ws.cell(row=row, column=1+c_idx, value=val)
            cell.border = thin_border
            # Nếu bị cảnh báo thì tô đỏ toàn bộ dòng hoặc chỉ cột trạng thái
            if is_warning:
                cell.font = red_font

    return send_excel_response(wb, f"CanhBao_{lop.id}")

def send_excel_response(wb, filename):
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={'Content-Disposition': f'attachment; filename="{filename}.xlsx"'})
    
    
# --- API TRANG ĐỆM HTML CHO XUẤT EXCEL ---
@router.get("/export/browser/{report_type}/{item_id}")
async def browser_download_wrapper(report_type: str, item_id: str):
    """
    Trang HTML đệm xử lý UX khi tải file, chống lỗi trình duyệt chặn pop-up/auto-download.
    """
    # Đường dẫn trỏ tới API xuất Excel thật
    real_download_url = f"/api/export/report/{report_type}/{item_id}"
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="vi">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Tải báo cáo AuEdu</title>
        <style>
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                display: flex; justify-content: center; align-items: center; 
                height: 100vh; background-color: #f8fafc; margin: 0;
            }}
            .card {{
                text-align: center; padding: 40px; background: white; 
                border-radius: 16px; box-shadow: 0 10px 25px rgba(0,0,0,0.05);
                border: 1px solid #e2e8f0; max-width: 400px; width: 90%;
            }}
            .spinner {{
                border: 4px solid #f1f5f9; border-top: 4px solid #3b82f6; 
                border-radius: 50%; width: 45px; height: 45px; 
                animation: spin 1s linear infinite; margin: 0 auto 20px auto;
            }}
            @keyframes spin {{ 0% {{ transform: rotate(0deg); }} 100% {{ transform: rotate(360deg); }} }}
            
            .success-icon {{
                display: none; font-size: 45px; color: #10b981; margin-bottom: 15px;
            }}
            h2 {{ color: #1e293b; margin-bottom: 10px; font-size: 20px; transition: 0.3s; }}
            p {{ color: #64748b; font-size: 14px; line-height: 1.5; margin-bottom: 25px; transition: 0.3s; }}
            
            .btn-download {{
                display: none; /* Ẩn lúc mới vào */
                background-color: #3b82f6; color: white; text-decoration: none;
                padding: 12px 24px; border-radius: 8px; font-weight: 600; font-size: 14px;
                transition: all 0.2s; border: none; cursor: pointer; box-shadow: 0 4px 6px rgba(59, 130, 246, 0.2);
            }}
            .btn-download:hover {{ background-color: #2563eb; transform: translateY(-1px); }}
            
            .brand {{ margin-top: 30px; font-size: 12px; color: #94a3b8; font-weight: 600; letter-spacing: 1px; }}
        </style>
    </head>
    <body>
        <div class="card">
            <div class="spinner" id="spinner"></div>
            
            <div class="success-icon" id="success-icon">✓</div>
            
            <h2 id="title-text">Đang chuẩn bị tệp Excel...</h2>
            <p id="desc-text">Trình duyệt sẽ tự động tải tệp báo cáo của bạn xuống trong giây lát.</p>
            
            <a href="{real_download_url}" class="btn-download" id="manualBtn">Tải xuống thủ công</a>
            
            <div class="brand">AUEDU ANALYTICS</div>
        </div>

        <script>
            // 1. Tự động kích hoạt tải xuống sau 1.5 giây
            setTimeout(function() {{
                window.location.href = "{real_download_url}";
                
                // 2. Thay đổi giao diện báo hiệu đã xử lý xong
                document.getElementById('spinner').style.display = 'none';
                document.getElementById('success-icon').style.display = 'block';
                
                document.getElementById('title-text').innerText = 'Đã hoàn tất yêu cầu!';
                document.getElementById('title-text').style.color = '#10b981';
                
                document.getElementById('desc-text').innerText = 'Nếu trình duyệt của bạn chặn tự động tải xuống, vui lòng nhấn nút bên dưới.';
                
                // Hiện nút tải lại thủ công
                document.getElementById('manualBtn').style.display = 'inline-block';
                
            }}, 1500);
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)