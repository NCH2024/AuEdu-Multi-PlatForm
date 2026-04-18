# server/app/api/reports.py
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, distinct
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

from app.db.session import get_db
from app.db.models import (
    ThoiKhoaBieu, Lop, HocPhan, LoaiHocPhan, HocKy, GiangVien,
    DiemDanh, TKBTiet, SinhVien,
    FaceEmbedding,
)

router = APIRouter()

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))


# -------------------------------------------------
# Helper: trả về file excel dưới dạng stream
# -------------------------------------------------
def stream_excel(wb, filename: str):
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content‑Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


# -------------------------------------------------
# 1️⃣ Báo cáo chi tiết (theo lớp‑tiết)
# -------------------------------------------------
@router.get("/export/report/detailed/{tkb_id}")
async def export_detailed_attendance(tkb_id: int, db: AsyncSession = Depends(get_db)):
    # Lấy thông tin TKB + header
    hdr_stmt = (
        select(ThoiKhoaBieu, Lop, HocPhan, HocKy, GiangVien)
        .join(Lop, ThoiKhoaBieu.lop_id == Lop.id)
        .join(HocPhan, ThoiKhoaBieu.hocphan_id == HocPhan.id)
        .join(HocKy, ThoiKhoaBieu.hocky_id == HocKy.id)
        .join(GiangVien, ThoiKhoaBieu.giangvien_id == GiangVien.id)
        .where(ThoiKhoaBieu.id == tkb_id)
    )
    res = await db.execute(hdr_stmt)
    first = res.first()
    if not first:
        raise HTTPException(status_code=404, detail="Không tìm thấy lịch học")
    tkb, lop, hp, hk, gv = first

    wb = load_workbook("app/templates/temp_DetailedAttendance.xlsx")
    ws = wb.active
    ws["C5"], ws["C6"], ws["C7"], ws["C8"], ws["C9"] = (
        lop.tenlop,
        hp.tenhocphan,
        hk.tenhocky,
        hk.namhoc,
        f"{gv.hodem} {gv.ten}",
    )

    # 2️⃣ Lấy danh sách ngày đã điểm danh
    date_stmt = (
        select(DiemDanh.ngay_diem_danh)
        .where(DiemDanh.tkb_tiet_id.in_(select(TKBTiet.id).where(TKBTiet.tkb_id == tkb_id)))
        .distinct()
        .order_by(DiemDanh.ngay_diem_danh.asc())
    )
    dates = (await db.execute(date_stmt)).scalars().all()
    for i, d in enumerate(dates):
        cell = ws.cell(row=12, column=5 + i, value=d.strftime("%d/%m"))
        cell.font, cell.border, cell.alignment = Font(bold=True), thin_border, Alignment(horizontal="center")

    # 3️⃣ Danh sách sinh viên lớp
    sv_stmt = select(SinhVien).where(SinhVien.class_id == lop.id).order_by(SinhVien.id.asc())
    students = (await db.execute(sv_stmt)).scalars().all()

    for r_idx, sv in enumerate(students):
        row = 13 + r_idx
        ws.cell(row=row, column=1, value=r_idx + 1).border = thin_border
        ws.cell(row=row, column=2, value=sv.id).border = thin_border
        ws.cell(row=row, column=3, value=sv.hodem).border = thin_border
        ws.cell(row=row, column=4, value=sv.ten).border = thin_border

        # Điểm danh của sinh viên này
        dd_stmt = select(DiemDanh).where(
            DiemDanh.sv_id == sv.id,
            DiemDanh.tkb_tiet_id.in_(
                select(TKBTiet.id).where(TKBTiet.tkb_id == tkb_id)
            ),
        )
        dd_map = {d.ngay_diem_danh: d.trang_thai for d in (await db.execute(dd_stmt)).scalars().all()}

        for c_idx, d in enumerate(dates):
            status = dd_map.get(d, "")
            val = "X" if status == "Có mặt" else ("V" if status == "Vắng" else "")
            cell = ws.cell(row=row, column=5 + c_idx, value=val)
            cell.border, cell.alignment = thin_border, Alignment(horizontal="center")

    return stream_excel(wb, f"ChiTietDiemDanh_{lop.id}")


# -------------------------------------------------
# 2️⃣ Báo cáo tổng quan (theo GV)
# -------------------------------------------------
@router.get("/export/report/overview/{gv_id}")
async def export_overview_report(gv_id: int, db: AsyncSession = Depends(get_db)):
    gv = await db.scalar(select(GiangVien).where(GiangVien.id == gv_id))
    if not gv:
        raise HTTPException(status_code=404, detail="Giảng viên không tồn tại")

    wb = load_workbook("app/templates/temp_OverviewReport.xlsx")
    ws = wb.active
    ws["C7"] = f"{gv.hodem} {gv.ten}" if gv else "N/A"

    # Lấy lớp, số sinh viên và số buổi đã dạy
    stmt = (
        select(Lop.id, Lop.tenlop, func.count(SinhVien.id).label("siso"))
        .join(ThoiKhoaBieu, Lop.id == ThoiKhoaBieu.lop_id)
        .join(SinhVien, Lop.id == SinhVien.class_id)
        .where(ThoiKhoaBieu.giangvien_id == gv_id)
        .group_by(Lop.id, Lop.tenlop)
    )
    data = (await db.execute(stmt)).all()

    for r_idx, item in enumerate(data):
        row = 10 + r_idx
        # Tính số buổi đã dạy, số lượt có mặt, tỷ lệ chuyên cần
        tkb_ids_sub = select(TKBTiet.id).join(ThoiKhoaBieu).where(ThoiKhoaBieu.lop_id == item.id)
        buoi_day = await db.scalar(
            select(func.count(func.distinct(DiemDanh.ngay_diem_danh))).where(
                DiemDanh.tkb_tiet_id.in_(tkb_ids_sub)
            ))or 0
        tong_co_mat = await db.scalar(
            select(func.count(DiemDanh.id)).where(DiemDanh.tkb_tiet_id.in_(tkb_ids_sub))
        ) or 0

        ty_le = (tong_co_mat / (item.siso * buoi_day) * 100) if (item.siso * buoi_day) > 0 else 0

        values = [r_idx + 1, item.id, item.tenlop, item.siso, buoi_day, f"{ty_le:.1f}%"]
        for c_idx, val in enumerate(values):
            cell = ws.cell(row=row, column=1 + c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    return stream_excel(wb, f"TongQuanLop_GV_{gv_id}")


# -------------------------------------------------
# 3️⃣ Báo cáo cảnh báo (sinh viên vắng quá mức)
# -------------------------------------------------
@router.get("/export/report/warning/{tkb_id}")
async def export_warning_report(tkb_id: int, db: AsyncSession = Depends(get_db)):
    # Thông tin TKB + Loại học phần
    stmt_hdr = (
        select(ThoiKhoaBieu, Lop, HocPhan, LoaiHocPhan, HocKy, GiangVien)
        .join(Lop, ThoiKhoaBieu.lop_id == Lop.id)
        .join(HocPhan, ThoiKhoaBieu.hocphan_id == HocPhan.id)
        .join(LoaiHocPhan, HocPhan.loaihp_id == LoaiHocPhan.id)
        .join(HocKy, ThoiKhoaBieu.hocky_id == HocKy.id)
        .join(GiangVien, ThoiKhoaBieu.giangvien_id == GiangVien.id)
        .where(ThoiKhoaBieu.id == tkb_id)
    )
    res = await db.execute(stmt_hdr)
    first = res.first()
    if not first:
        raise HTTPException(status_code=404, detail="Không tìm thấy lịch học")
    tkb, lop, hp, lhp, hk, gv = first

    # Ngưỡng cảnh báo (theo loại học phần)
    threshold = 0.3 if "Lý thuyết" in lhp.tenloai else 0.2
    max_absences = hp.sobuoi * threshold

    # Số buổi đã dạy
    tiết_ids = select(TKBTiet.id).where(TKBTiet.tkb_id == tkb_id)
    total_sessions = await db.scalar(
        select(func.count(func.distinct(DiemDanh.ngay_diem_danh))).where(
            DiemDanh.tkb_tiet_id.in_(tiết_ids))
    ) or 0

    wb = load_workbook("app/templates/temp_WarningList.xlsx")
    ws = wb.active
    ws["C5"], ws["C6"], ws["C7"], ws["C8"], ws["C9"] = (
        lop.tenlop,
        hp.tenhocphan,
        hk.tenhocky,
        hk.namhoc,
        f"{gv.hodem} {gv.ten}",
    )

    # Danh sách sinh viên lớp
    sv_stmt = select(SinhVien).where(SinhVien.class_id == lop.id).order_by(SinhVien.id.asc())
    students = (await db.execute(sv_stmt)).scalars().all()

    red_font = Font(color="FF0000", bold=True)

    for r_idx, sv in enumerate(students):
        row = 12 + r_idx
        # Số buổi sinh viên có mặt
        attended = await db.scalar(
            select(func.count(DiemDanh.id)).where(
                DiemDanh.sv_id == sv.id,
                DiemDanh.trang_thai == "Có mặt",
                DiemDanh.tkb_tiet_id.in_(tiết_ids),
            )
        ) or 0
        absent = total_sessions - attended
        is_warning = absent > max_absences
        percent = (absent / hp.sobuoi * 100) if hp.sobuoi > 0 else 0

        vals = [
            r_idx + 1,
            sv.id,
            sv.hodem,
            sv.ten,
            attended,
            absent,
            f"{percent:.1f}%",
            "CẢNH BÁO" if is_warning else "Bình thường",
        ]
        for c_idx, val in enumerate(vals):
            cell = ws.cell(row=row, column=1 + c_idx, value=val)
            cell.border = thin_border
            if is_warning:
                cell.font = red_font
            # Căn giữa cho một số cột
            if c_idx >= 4:
                cell.alignment = Alignment(horizontal="center")

    return stream_excel(wb, f"CanhBao_{lop.id}")


# -------------------------------------------------
# 4️⃣ HTML wrapper để kích hoạt download (tránh chặn pop‑up)
# -------------------------------------------------
@router.get("/export/browser/{report_type}/{item_id}")
async def browser_download_wrapper(report_type: str, item_id: str):
    """
    Trang HTML đệm xử lý UX khi tải file, chống lỗi trình duyệt chặn pop-up/auto-download.
    """
    # Đường dẫn trỏ tới API xuất Excel thật
    real_download_url = f"/api/export/report/{report_type}/{item_id}"
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="vi">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Tải báo cáo AuEdu</title>
        <style>
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                display: flex; justify-content: center; align-items: center; 
                height: 100vh; background-color: #f8fafc; margin: 0;
            }}
            .card {{
                text-align: center; padding: 40px; background: white; 
                border-radius: 16px; box-shadow: 0 10px 25px rgba(0,0,0,0.05);
                border: 1px solid #e2e8f0; max-width: 400px; width: 90%;
            }}
            .spinner {{
                border: 4px solid #f1f5f9; border-top: 4px solid #3b82f6; 
                border-radius: 50%; width: 45px; height: 45px; 
                animation: spin 1s linear infinite; margin: 0 auto 20px auto;
            }}
            @keyframes spin {{ 0% {{ transform: rotate(0deg); }} 100% {{ transform: rotate(360deg); }} }}
            
            .success-icon {{
                display: none; font-size: 45px; color: #10b981; margin-bottom: 15px;
            }}
            h2 {{ color: #1e293b; margin-bottom: 10px; font-size: 20px; transition: 0.3s; }}
            p {{ color: #64748b; font-size: 14px; line-height: 1.5; margin-bottom: 25px; transition: 0.3s; }}
            
            .btn-download {{
                display: none; /* Ẩn lúc mới vào */
                background-color: #3b82f6; color: white; text-decoration: none;
                padding: 12px 24px; border-radius: 8px; font-weight: 600; font-size: 14px;
                transition: all 0.2s; border: none; cursor: pointer; box-shadow: 0 4px 6px rgba(59, 130, 246, 0.2);
            }}
            .btn-download:hover {{ background-color: #2563eb; transform: translateY(-1px); }}
            
            .brand {{ margin-top: 30px; font-size: 12px; color: #94a3b8; font-weight: 600; letter-spacing: 1px; }}
        </style>
    </head>
    <body>
        <div class="card">
            <div class="spinner" id="spinner"></div>
            
            <div class="success-icon" id="success-icon">✓</div>
            
            <h2 id="title-text">Đang chuẩn bị tệp Excel...</h2>
            <p id="desc-text">Trình duyệt sẽ tự động tải tệp báo cáo của bạn xuống trong giây lát.</p>
            
            <a href="{real_download_url}" class="btn-download" id="manualBtn">Tải xuống thủ công</a>
            
            <div class="brand">AUEDU ANALYTICS</div>
        </div>

        <script>
            // 1. Tự động kích hoạt tải xuống sau 1.5 giây
            setTimeout(function() {{
                window.location.href = "{real_download_url}";
                
                // 2. Thay đổi giao diện báo hiệu đã xử lý xong
                document.getElementById('spinner').style.display = 'none';
                document.getElementById('success-icon').style.display = 'block';
                
                document.getElementById('title-text').innerText = 'Đã hoàn tất yêu cầu!';
                document.getElementById('title-text').style.color = '#10b981';
                
                document.getElementById('desc-text').innerText = 'Nếu trình duyệt của bạn chặn tự động tải xuống, vui lòng nhấn nút bên dưới.';
                
                // Hiện nút tải lại thủ công
                document.getElementById('manualBtn').style.display = 'inline-block';
                
            }}, 1500);
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)